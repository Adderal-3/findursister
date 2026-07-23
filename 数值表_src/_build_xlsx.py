# -*- coding: utf-8 -*-
"""构建《找你妹大作战》数值表.xlsx v2.5：封面 + 5 张数值表 + 数值决议（D1~D27）"""
import csv, os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import LineChart, Reference
from openpyxl.utils import get_column_letter

SRC = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(SRC)
OUT = os.path.join(ROOT, '数值表.xlsx')

# ---------- 样式（极简黑白灰 + 蓝色点缀） ----------
F_TITLE = Font(size=18, bold=True, color='000000')
F_SUB = Font(size=11, color='666666')
F_HEAD = Font(size=10, bold=True, color='FFFFFF')
F_INPUT = Font(size=10, color='0066CC')   # 静态输入 = 蓝
F_CALC = Font(size=10, color='000000')    # 公式 = 黑
F_NOTE = Font(size=9, color='999999')
FILL_HEAD = PatternFill('solid', start_color='333333')
FILL_BAND = PatternFill('solid', start_color='F5F5F5')
FILL_KEY = PatternFill('solid', start_color='E6F0FA')
CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')

def sheet_head(ws, title, subtitle, headers, widths, start_col=2):
    ws.sheet_view.showGridLines = False
    r = 2
    ws.merge_cells(start_row=r, start_column=start_col, end_row=r, end_column=start_col + len(headers) - 1)
    c = ws.cell(r, start_col, title); c.font = F_TITLE
    ws.row_dimensions[r].height = 30
    r += 1
    ws.merge_cells(start_row=r, start_column=start_col, end_row=r, end_column=start_col + len(headers) - 1)
    c = ws.cell(r, start_col, subtitle); c.font = F_SUB
    r += 1
    for j, h in enumerate(headers):
        c = ws.cell(r, start_col + j, h); c.font = F_HEAD; c.fill = FILL_HEAD; c.alignment = CENTER
    for j, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(start_col + j)].width = w
    return r + 1  # 数据起始行

def put(ws, r, c, v, font=F_INPUT, align=CENTER):
    cell = ws.cell(r, c, v); cell.font = font; cell.alignment = align
    return cell

wb = Workbook()

# ================= 1. 封面 =================
ws = wb.active; ws.title = '封面'
ws.sheet_view.showGridLines = False
for col, w in zip('BCDEFGH', (14, 16, 16, 16, 16, 16, 16)):
    ws.column_dimensions[col].width = w
ws.merge_cells('B2:H2'); put(ws, 2, 2, '《找你妹大作战》数值表 v2.5', F_TITLE, LEFT); ws.row_dimensions[2].height = 34
ws.merge_cells('B3:H3'); put(ws, 3, 2, '配套《玩法设计文档 v2.5》· 谜字坊规则本地化 · 供关卡 / 系统 / 数值开发使用', F_SUB, LEFT)
ws.merge_cells('B5:H5'); put(ws, 5, 2, '核心规模', Font(size=13, bold=True), LEFT)
metrics = [('主线关卡', '100 关（5 篇章 / 10 小章）'), ('伙伴', '8 位 × 8 种招募方式'), ('物品素材', '6 主题 × 40 = 240 件'),
           ('道具', '4 种（无尽模式禁用）'), ('无尽波次表', '40 波（含 3 个高压波）'), ('理论满加成', '+22%（图鉴12+伙伴10）'),
           ('星级锚模型', 'StarBase = T×M(n) − 0.4×h(n)（极限速点，D10）')]
r = 6
for k, v in metrics:
    put(ws, r, 2, k, Font(size=10, bold=True), LEFT).fill = FILL_KEY
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    put(ws, r, 3, v, F_CALC, LEFT); r += 1
r += 1
put(ws, r, 2, '工作表索引', Font(size=13, bold=True), LEFT); r += 1
idx = [('关卡参数表', 'P0 · 100 关逐关参数（类型/目标/干扰/时长/星级锚/星级线），星级列为公式'),
       ('星锚系数表', 'P1 · M(n)/h(n) 查表（n=1..60），关卡表星级锚的数据源'),
       ('无尽难度曲线', 'P2 · 第 1~40 波参数与预期清关耗时'),
       ('道具经济表', 'P2 · 4 道具的获取/消耗（分无视频/看视频玩家）/上限/反滥用'),
       ('伙伴招募条件', 'P1 · 8 伙伴招募条件速查（完整规格见 数值表_src/partners_spec.md）'),
       ('数值决议', '全组统一的数值口径决议 D1~D27（开发必读）')]
for name, desc in idx:
    put(ws, r, 2, name, Font(size=10, bold=True, color='0066CC'), LEFT)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=8)
    put(ws, r, 3, desc, F_CALC, LEFT); r += 1
r += 1
ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)
put(ws, r, 2, '静态蓝字 = 输入参数；黑字 = 公式计算。源数据与生成脚本见 数值表_src/ 目录。', F_NOTE, LEFT)

# ================= 2. 关卡参数表 =================
ws = wb.create_sheet('关卡参数表')
heads = ['关卡', '小章', '主题', '类型', '目标种类', '每种份数', '目标总数', '干扰项', '类型系数', '限时(秒)', '星级锚', '⭐⭐线', '⭐⭐⭐线', '解锁需星']
wids = [7, 6, 11, 11, 9, 9, 9, 8, 9, 9, 11, 10, 10, 9]
row0 = sheet_head(ws, '关卡参数表（100 关）', '蓝字=输入参数；黑字=公式。类型：standard标准 cluster集簇 mist迷踪 night夜航 speed极速 boss。星级锚=T×M(n)−0.4×h(n)，M/h 见《星锚系数表》（D10）', heads, wids)
with open(os.path.join(SRC, 'levels_100.csv'), encoding='utf-8') as f:
    rows = list(csv.DictReader(f))
for i, d in enumerate(rows):
    r = row0 + i
    put(ws, r, 2, int(d['level'])); put(ws, r, 3, int(d['chapter'])); put(ws, r, 4, d['theme'])
    put(ws, r, 5, d['type']); put(ws, r, 6, int(d['targetTypes'])); put(ws, r, 7, int(d['copiesPerTarget']))
    put(ws, r, 8, f'=F{r}*G{r}', F_CALC)
    put(ws, r, 9, int(d['distractors'])); put(ws, r, 10, float(d['typeCoef']))
    put(ws, r, 11, f'=ROUND((45+H{r}*8)*J{r},0)', F_CALC)
    put(ws, r, 12, (f'=ROUND(K{r}*(IF(H{r}<=10,0.9*H{r}+0.05*H{r}*(H{r}+1),14.5+1.9*(H{r}-10)))'
                    f'-0.4*IF(H{r}<=10,0.45*H{r}*(H{r}-1)+H{r}*(H{r}^2-1)/30,0.95*H{r}*(H{r}-1)-12),2)'), F_CALC)
    put(ws, r, 13, f'=FLOOR(L{r}*0.55,0.1)', F_CALC)
    put(ws, r, 14, f'=FLOOR(L{r}*0.75,0.1)', F_CALC)
    if d['starUnlockReq']:
        put(ws, r, 15, int(d['starUnlockReq']))
    if d['type'] == 'boss':
        for c in range(2, 16):
            ws.cell(r, c).fill = FILL_KEY
    elif i % 2:
        for c in range(2, 16):
            ws.cell(r, c).fill = FILL_BAND
ws.freeze_panes = f'A{row0}'
# 图表：星级锚曲线
ch = LineChart(); ch.title = '各关星级锚曲线'; ch.style = 2; ch.height = 7; ch.width = 22
data = Reference(ws, min_col=12, min_row=row0 - 1, max_row=row0 + 99)
cats = Reference(ws, min_col=2, min_row=row0, max_row=row0 + 99)
ch.add_data(data, titles_from_data=True); ch.set_categories(cats); ch.legend = None
ws.add_chart(ch, f'B{row0 + 102}')

# ================= 3. 星锚系数表 =================
ws = wb.create_sheet('星锚系数表')
heads = ['目标总数 n', 'M(n)', 'h(n)']
wids = [11, 13, 13]
row0 = sheet_head(ws, '星锚系数表 M(n) / h(n)',
                'StarBase = T × M(n) − 0.4 × h(n)（D10 极限速点模型）。M：n≤10 → 0.9n+0.05n(n+1)；n≥10 → 14.5+1.9(n−10)。h：n≤10 → 0.45n(n−1)+n(n²−1)/30；n≥10 → 0.95n(n−1)−12', heads, wids)
for n in range(1, 61):
    r = row0 + n - 1
    put(ws, r, 2, n)
    put(ws, r, 3, f'=IF(B{r}<=10,0.9*B{r}+0.05*B{r}*(B{r}+1),14.5+1.9*(B{r}-10))', F_CALC)
    put(ws, r, 4, f'=IF(B{r}<=10,0.45*B{r}*(B{r}-1)+B{r}*(B{r}^2-1)/30,0.95*B{r}*(B{r}-1)-12)', F_CALC)
    ws.cell(r, 3).number_format = '0.000000'
    ws.cell(r, 4).number_format = '0.000000'
ws.freeze_panes = f'A{row0}'

# ================= 4. 无尽难度曲线 =================
ws = wb.create_sheet('无尽难度曲线')
heads = ['波次', '目标种类', '每种份数', '目标总数', '干扰项', '跨主题', '类型系数', '参考时长(秒)', '预期清关(秒)']
wids = [7, 9, 9, 9, 8, 8, 9, 11, 11]
row0 = sheet_head(ws, '无尽模式难度曲线（第 1~40 波）', '第 15/25/35 波为高压波（蓝底）。参考时长仅为对账口径，实际无尽连续计时不清零；预期清关为模型估算值', heads, wids)
with open(os.path.join(SRC, 'endless_curve.csv'), encoding='utf-8') as f:
    rows = list(csv.DictReader(f))
for i, d in enumerate(rows):
    r = row0 + i
    put(ws, r, 2, int(d['wave'])); put(ws, r, 3, int(d['targetTypes'])); put(ws, r, 4, int(d['copiesPerTarget']))
    put(ws, r, 5, f'=C{r}*D{r}', F_CALC)
    put(ws, r, 6, int(d['distractors'])); put(ws, r, 7, int(d['crossTheme'])); put(ws, r, 8, float(d['typeCoef']))
    put(ws, r, 9, f'=ROUND((45+E{r}*8)*H{r},0)', F_CALC)
    put(ws, r, 10, float(d['expectedClearSec']), F_INPUT)
    if int(d['wave']) in (15, 25, 35):
        for c in range(2, 11):
            ws.cell(r, c).fill = FILL_KEY
    elif i % 2:
        for c in range(2, 11):
            ws.cell(r, c).fill = FILL_BAND
ws.freeze_panes = f'A{row0}'
ch = LineChart(); ch.title = '干扰项数量随波次'; ch.style = 2; ch.height = 7; ch.width = 20
data = Reference(ws, min_col=6, min_row=row0 - 1, max_row=row0 + 39)
cats = Reference(ws, min_col=2, min_row=row0, max_row=row0 + 39)
ch.add_data(data, titles_from_data=True); ch.set_categories(cats); ch.legend = None
ws.add_chart(ch, f'B{row0 + 42}')

# ================= 5. 道具经济表 =================
ws = wb.create_sheet('道具经济表')
heads = ['道具', '效果', '每关上限', '获取途径', '日获取(无视频)', '日获取(看视频)', '日消耗(无视频)', '日消耗(看视频)', '存量上限', '反滥用规则']
wids = [10, 26, 16, 40, 11, 11, 11, 11, 9, 42]
row0 = sheet_head(ws, '道具经济平衡表', '测算口径：无视频玩家中心 40 关/日；看视频玩家中心 65 关/日（模型 55~75，运营目标 60~75）。全部道具无尽禁用；视频规则见 D16/D23/D24', heads, wids)
with open(os.path.join(SRC, 'props_economy.csv'), encoding='utf-8') as f:
    rows = list(csv.reader(f))
heads_csv, data_rows = rows[0], rows[1:]
for i, d in enumerate(data_rows):
    r = row0 + i
    for j, v in enumerate(d):
        try:
            v2 = float(v)
        except ValueError:
            v2 = v
        put(ws, r, 2 + j, v2, F_INPUT, LEFT if j in (1, 3, 9) else CENTER)
    if i % 2:
        for c in range(2, 12):
            ws.cell(r, c).fill = FILL_BAND

# ================= 6. 伙伴招募条件 =================
ws = wb.create_sheet('伙伴招募条件')
heads = ['#', '伙伴', '主题', '招募方式', '条件（判定口径）', '指标 metricKey', '类型', '进度显示格式']
wids = [5, 15, 9, 10, 34, 26, 8, 18]
row0 = sheet_head(ws, '伙伴招募条件速查（8 位 × 8 种方式）', '每位 +1.25%、满 8 人 +10%，仅终身榜/赛季榜/无尽榜生效，日榜不吃加成。阈值已按 v2.4 重标（D17）。完整规格卡/文案/边界细则见 数值表_src/partners_spec.md', heads, wids)
partners = [
    (1, '🍳 小厨神·阿福', 'kitchen', '在线时长', '累计在线 ≥ 1800 秒（30 分钟）', 'znm.prog.onlineSec', '累计', '在线 24/30 分钟'),
    (2, '🧸 玩具兵·锡锡', 'toys', '活跃', '单日累计游玩 ≥ 1200 秒（20 分钟）任意一天（D15：不绑连续签到）', 'znm.prog.dailyPlaySec', '单日', '单日游玩 14/20 分钟'),
    (3, '🦉 林间向导·咕咕', 'nature', '寻物', 'nature 主题图鉴点亮 ≥ 15 件', 'znm.coll.theme.nature', '累计', '自然图鉴 11/15'),
    (4, '🧭 旅行家·远远', 'travel', '任务', '历史最高星级为三星的关卡累计 ≥ 10 关（只增不减）', 'znm.best.levelStars', '累计', '三星通关 7/10'),
    (5, '🐬 潜水员·泡泡', 'ocean', '积累', '累计命中目标 ≥ 500 次（含重复物品）', 'znm.prog.totalFinds', '累计', '累计找物 163/500'),
    (6, '🚀 宇航员·星星', 'space', '分数', '单关基础分 ≥ 1500（不含图鉴/伙伴系数）', 'znm.prog.bestBaseScore', '单次', '单关最高 1286/1500'),
    (7, '🔍 侦探·眯眯', '彩蛋', '技巧', '单局最大连击 ≥ 8（任意模式）', 'znm.prog.maxCombo', '单次', '最大连击 6/8'),
    (8, '👧 妹妹本妹', '终极彩蛋', '羁绊', '其余 7 位全部招募后自动加入（不占常规判定）', 'znm.partner.roster', '事件', '伙伴 5/7'),
]
for i, p in enumerate(partners):
    r = row0 + i
    for j, v in enumerate(p):
        put(ws, r, 2 + j, v, F_INPUT, LEFT if j in (1, 4, 5, 7) else CENTER)
    if i % 2:
        for c in range(2, 10):
            ws.cell(r, c).fill = FILL_BAND

# ================= 7. 数值决议 =================
ws = wb.create_sheet('数值决议')
heads = ['编号', '决议', '背景与影响']
wids = [7, 52, 60]
row0 = sheet_head(ws, '数值口径决议 D1~D27（开发必读）', 'D1~D8 为基础口径；D9~D20 为 v2.4 修复；D21~D27 为 v2.5 开发级收口', heads, wids)
decisions = [
    ('D1', '连击倍率上限 = 1.9（10 连击封顶），不是 2.0', '公式 1+0.1×(c−1) 在 c=10 时为 1.9'),
    ('D2', '关卡类型占比：标准 55 / 集簇 10 / 迷踪 10 / 夜航 10 / 极速 5 / Boss 10', '原契约合计 105 超限，标准关取余量 55，合计恰好 100'),
    ('D3', '舍入统一 half-up（四舍五入）', '限时 T 的 round 与星级锚均按 half-up；Python 侧用 Decimal，避免银行家舍入差 1 秒'),
    ('D4', '星级线 = FLOOR(星级锚 × 55% / 75%, 0.1)', '向下取整到 0.1；注意浮点陷阱，实现时 +1e-6 再 floor'),
    ('D5', '终身榜/赛季榜存基础分，伙伴系数读取时按当前人数实时重算', '招募新伙伴后历史成绩立即吃新系数，无需重玩；日榜永不吃伙伴系数'),
    ('D6', '无尽模式禁用全部 4 种道具', '保住无尽榜公平性；道具消耗模型已按此建立'),
    ('D7', 'Boss 关与无尽高压波同屏物品可达 ~163/115，超出原 ~110 容量注释', '引擎侧需实测：必要时上调 SCENE_SCALE 或下调 minGap；数值表暂按契约原样输出'),
    ('D8', '伙伴判定检查点 = 结算 / 签到 / 图鉴点亮 / onlineTick(30s) / 启动兜底', '纯挂机玩家也能按在线时长招到 1 号阿福'),
    ('D9', '连击无时间窗口，仅点错清零；无尽跨波保留', 'v2.4 修复：删除原 3.5s 窗口；记忆背板+快速连点是合法高手打法'),
    ('D10', '星级锚 = StarBase = T×M(n) − 0.4×h(n)（极限速点模型）', '替代旧"理论满分"口径；锚点：n=12,T=141 → 2534.94；n=36,T=300 → 18696.00'),
    ('D11', '体力在线每 3 分钟 +1；体力视频单次 +5', '完整约束见 D24'),
    ('D12', '图鉴全局平铺：每主题 +2% 全局（满 +12%）；日榜不吃图鉴', '替代按主题加成；消除偏科'),
    ('D13', '日榜 = 当日各关最高基础分之和，每关每日只计一次，不吃任何加成', '钉死口径，防止系数污染日榜'),
    ('D14', '终身/赛季总榜同分按对应基础分达到时间，系数变化不刷新时间戳', '防止后招募伙伴者靠系数偷排名'),
    ('D15', '无掉队设计：永久进度不绑自然日累计；每日上限仅限消耗品', '对用户核心诉求的硬约束；锡锡条件由"累计签到2天"改为"单日游玩20分钟"'),
    ('D16', '站内视频三入口：体力 +5 / 局中加时固定 +15s / 卡关放大镜 +1', '导流主通道；加时与沙漏共享每关≤3 次上限'),
    ('D17', '伙伴阈值重标：阿福 1800s / 锡锡单日 1200s / 咕咕 15 件 / 远远 10 三星 / 泡泡 500 件 / 星星 1500 分 / 眯眯 8 连击 / 妹妹 7 人齐', '对齐新计分量级（旧 200 命中/300 分在新模型下失真）'),
    ('D18', 'Boss 关类型系数 0.9（L10=213s、L20 起=300s）', '替代旧 1.3，告别 7.2 分钟疲劳局'),
    ('D19', 'UGC 隔离：1 眼力/次，成绩不计榜单/招募/图鉴', '时长30~300s、物品≤130、目标≤已摆放实例'),
    ('D20', '时光沙漏 = 免看视频固定 +15s，与视频加时共享每关≤3次', '两个入口满格时同时置灰'),
    ('D21', '层级 = 5篇章×20关 / 10小章×10关', 'CSV chapter 指小章；星门解锁篇章'),
    ('D22', '剩余时间下截0.1s → 单物品得分 half-up 0.1 → 逐项求和', '压线事件按服务端时间排序'),
    ('D23', '视频期间暂停计时/输入，完播验签后幂等发奖', '加时固定+15s；rewardId 防重复'),
    ('D24', '体力上限/初始20；每日补到20；视频+5每日≤4次', '所有恢复不溢出'),
    ('D25', '终身总榜不重置；赛季榜独立记录当季各关最好成绩', '赛季榜按赛季基础分达到时间破同分；奖励只绑定赛季榜'),
    ('D26', 'UGC奖励以另一有效账号首次完成作为条件', '作者/同设备/重复完成不计，日限3'),
    ('D27', '每主题28普通/8稀有/4史诗；首槽优先未点亮', '其余槽按70%/25%/5%权重'),
]
for i, (no, dec, bg) in enumerate(decisions):
    r = row0 + i
    put(ws, r, 2, no, Font(size=10, bold=True, color='0066CC'))
    put(ws, r, 3, dec, F_INPUT, LEFT)
    put(ws, r, 4, bg, F_CALC, LEFT)
    if i % 2:
        for c in range(2, 5):
            ws.cell(r, c).fill = FILL_BAND

wb.save(OUT)
print('saved:', OUT)
